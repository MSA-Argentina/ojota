"""
This file is part of Ojota.

    Ojota is free software: you can redistribute it and/or modify
    it under the terms of the GNU LESSER GENERAL PUBLIC LICENSE as published by
    the Free Software Foundation, either version 3 of the License, or
    (at your option) any later version.

    Ojota is distributed in the hope that it will be useful,
    but WITHOUT ANY WARRANTY; without even the implied warranty of
    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
    GNU Lesser General Public License for more details.

    You should have received a copy of the GNU  Lesser General Public License
    along with Ojota.  If not, see <http://www.gnu.org/licenses/>.
"""
from __future__ import absolute_import
from __future__ import print_function
import os
import json
from six.moves import zip

try:
    import yaml
    yaml_imported = True
except ImportError:
    yaml_imported = False

try:
    import requests
    request_imported = True
except ImportError:
    request_imported = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.cell import get_column_letter
    from openpyxl.shared.exc import InvalidFileException
    openpyxl_imported = True
except ImportError:
    openpyxl_imported = False

try:
    import dson
    dson_imported = True
except ImportError:
    dson_imported = False


_DATA_SOURCE = "data"


class Source(object):
    """Base class for all the data sources."""
    def __init__(self, data_path=None, create_empty=True):
        """Constructor for the Source class.

        Arguments:
        data_path -- the path where the data is located.
        """
        self.data_path = data_path
        self.create_empty = create_empty

    def _get_file_path(self, cls):
        """Builds the path where the data will be located.

        Arguments:
            cls -- the class with the data.
        """
        if self.data_path is None:
            data_path = _DATA_SOURCE
        else:
            data_path = self.data_path
        if cls.data_in_root or not cls.get_current_data_code():
            filepath = os.path.join(data_path, cls.get_plural_name())
        else:
            filepath = os.path.join(data_path, cls.get_current_data_code(),
                                    cls.get_plural_name())
        return filepath

    def fetch_elements(self, cls):
        """Fetch the elements for a given class.

        Arguments:
            cls - the class with the data.
        """
        data_path = self._get_file_path(cls)
        return self.read_elements(cls, data_path)

    def fetch_element(self, cls, pk):
        """Fetch the elements for a given element of a class.

        Arguments:
            cls - the class with the data.
            pk - the primary key of the given element.
        """
        data_path = self._get_file_path(cls)
        return self.read_element(cls, data_path, pk)

    def save(self, cls, data):
        """Fetch the elements for a given element of a class.

        Arguments:
            cls - the class with the data.
            pk - the primary key of the given element.
        """
        file_path = self._get_file_path(cls)
        self.write_elements(file_path, data)

    def read_elements(self, cls, filepath):
        raise NotImplementedError

    def read_element(self, cls, url, pk):
        raise NotImplementedError

    def write_elements(self, filepath, data):
        raise NotImplementedError


class JSONSource(Source):
    """Source class for the data stored with JSON format"""

    def __init__(self, data_path=None, create_empty=True, indent=4):
        """Constructor for the Source class.

        Arguments:
            data_path -- the path where the data is located.
            create_empty -- if file in data_path is not found, create an empty one.
            indent -- control the indentation of the JSON in the file.
        """
        self.indent = indent
        super(JSONSource, self).__init__(data_path, create_empty)

    def read_elements(self, cls, filepath):
        """Reads the elements form a JSON file. Returns a dictionary containing
        the read data.

        Arguments:
            filepath -- the path for the json file.
        """
        json_path = '%s.json' % filepath
        try:
            json_file = open(json_path, 'r')
            data = json.load(json_file)
            json_file.close()
        except IOError:
            if self.create_empty:
                json_file = open(json_path, 'w')
                json_file.write("[]")
                json_file.close()
                json_file = open(json_path, 'r')
                data = json.load(json_file)
                json_file.close()
            else:
                data = []
        try:
            elements = dict((element_data[cls.pk_field], element_data)
                            for element_data in data)
        except KeyError:
            msg = "Primary key was not found. Check that you have "
            msg += "configured the class correctly. In case you "
            msg += "have check your data source"
            raise AttributeError(msg)

        return elements

    def write_elements(self, filepath, data):
        data_set = open('%s.json' % filepath, 'w')
        json_data = json.dumps(data, indent=self.indent)
        data_set.write(json_data)
        data_set.close()


class YAMLSource(Source):
    """Source class for the data stored with YAML format.

    requires the PyYaml package to run.
    """
    def read_elements(self, cls, filepath):
        """Reads the elements form a JSON file. Returns a dictionary containing
        the read data.

        Arguments:
            filepath -- the path for the json file.
        """
        if yaml_imported:
            datos = yaml.load(open('%s.yaml' % filepath, 'r'))
            elements = {}
            for key, value in list(datos.items()):
                elements[value[cls.pk_field]] = value
        else:
            msg = "In order to use YAML sources you should install "
            msg += " the 'PyYAML' package"
            raise Exception(msg)

        return elements


class WebServiceSource(Source):
    """Source class for the data stored with JSON format taken through a Web
    Service.

        Requires the "requests" package to run.
        http://pypi.python.org/pypi/requests

    """
    WSTIMEOUT = 5

    def __init__(self, data_path=None, method="get", get_all_cmd="/all",
                 get_cmd="/data", user=None, password=None, cert=None,
                 custom_call=None):
        """Constructor for the WebServiceSource class.

        Arguments:
            data_path -- the path where the data is located.
            method -- the http method that will be used witht the web service.
            Defauts to "get".
            get_all_cmd -- the WS command to fetch all the data.
            Defaults to "/all".
            get_cmd -- the WS command to fetch one element.
            Defaults to "/data"
            user -- the user name for the authentication. If not provided
            the request will not use authentication.
            password -- the password for the authentication. If not
            provided the request will not use authentication.
        """
        Source.__init__(self, data_path=data_path)

        if request_imported:
            self.get_cmd = get_cmd
            self.get_all_cmd = get_all_cmd
            self.custom_call = custom_call if custom_call is not None else ""
            self.cert = cert
            try:
                self.method = getattr(requests, method)
            except AttributeError:
                self.method = requests.get
            if user is not None and password is not None:
                self.auth = (user, password)
            else:
                self.auth = None
        else:
            msg = "In order to use Web Service sources you should "
            msg += " install the 'requests' package"
            raise Exception(msg)

    def read_elements(self, cls, url):
        """Reads the elements form a WS request. Returns a dictionary
        containing the read data.

        Arguments:
            cls -- the data class.
            url -- the path for the WS.
        """
        _url = url + self.get_all_cmd
        verify = self.cert is not None
        response = self.method(_url, timeout=self.WSTIMEOUT, auth=self.auth,
                               cert=self.cert, verify=verify)
        data = response.json
        elements = dict((element_data[cls.pk_field], element_data)
                        for element_data in data)
        return elements

    def read_element(self, cls, url, pk):
        """Reads one element elements form a JSON file. Returns a dictionary
        containing the read data.

        Arguments:
            cls -- the data class.
            url -- the path for the WS.
            pk -- the primary key.
        """
        _url = "%s/%s%s" % (url, pk, self.get_cmd)
        verify = self.cert is not None
        response = self.method(_url, timeout=self.WSTIMEOUT, auth=self.auth,
                               cert=self.cert, verify=verify)
        data = response.json
        element = {data[cls.pk_field]: data}
        return element


class CSVSource(Source):
    def __init__(self, data_path=None, separator=","):
        Source.__init__(self, data_path=data_path)
        self.separator = separator

    """Source class for the data stored with JSON format"""
    def read_elements(self, cls, filepath):
        """Reads the elements form a JSON file. Returns a dictionary containing
        the read data.

        Arguments:
            filepath -- the path for the json file.
        """
        data = open('%s.csv' % filepath, 'r')
        keys = data.readline().strip().split(self.separator)
        dicts = [dict(list(zip(keys, elem.strip().split(
            self.separator)))) for elem in data]

        try:
            elements = {}
            for element in dicts:
                for key, value in list(element.items()):
                    if value == "":
                        del element[key]
                elements[element[cls.pk_field]] = element
        except KeyError:
            msg = "Primary key was not found. Check that you have "
            msg += "configured the class correctly. In case you "
            msg += "have check your data source"
            raise AttributeError(msg)

        return elements

    def write_elements(self, filepath, data):
        data_set = open('%s.csv' % filepath, 'w')
        keys = []
        for element in data:
            keys.extend(list(element.keys()))
        keys = set(keys)
        lines = []
        lines.append(self.separator.join(keys) + "\n")
        for element in data:
            line = self.separator.join([element.get(key, "") for key in keys])
            lines.append(line + "\n")

        data_set.writelines(lines)
        data_set.close()


class XLSSource(Source):
    def __init__(self, data_path=None, worksheet=0):
        Source.__init__(self, data_path=data_path)
        self.worksheet = worksheet

    """Source class for the data stored with JSON format"""
    def read_elements(self, cls, filepath):
        """Reads the elements form a JSON file. Returns a dictionary containing
        the read data.

        Arguments:
            filepath -- the path for the json file.
        """
        elements = {}
        if openpyxl_imported:
            try:
                wb = load_workbook('%s.xlsx' % filepath)
                ws = wb.get_active_sheet()
                rows = list(ws.rows)
                keys = [row.value for row in rows.pop(0)]
                for row in rows:
                    row_dict = {}
                    for key, cell in enumerate(row):
                        row_dict[keys[key]] = cell.value
                    elements[row_dict[cls.pk_field]] = row_dict
            except InvalidFileException:
                print("Warning, the file in invalid")
        else:
            raise Exception("In order to use XLS sources you should install the 'openpyxl' package")

        return elements

    def write_elements(self, filepath, data):
        wb = Workbook()
        dest_filename = '%s.xlsx' % filepath
        ws = wb.worksheets[self.worksheet]
        ws.title = "Ojota data"

        keys = []
        for element in data:
            keys.extend(list(element.keys()))
        keys = list(set(keys))
        row = 1
        for col_index, key in enumerate(keys, 1):
            col = get_column_letter(col_index)
            ws.cell('%s%s' % (col, row)).value = key

        for row, element in enumerate(data, 2):
            for col_index, key in enumerate(keys, 1):
                col = get_column_letter(col_index)
                ws.cell('%s%s' % (col, row)).value = element.get(key)

        wb.save(filename=dest_filename)


class DSONSource(Source):
    """Source class for the data stored with JSON format"""
    def read_elements(self, cls, filepath):
        """Reads the elements form a DSON file. Returns a dictionary containing
        the read data.

        Arguments:
            filepath -- the path for the dson file.
        """
        if dson_imported:
            dson_path = '%s.dson' % filepath
            try:
                dson_file = open(dson_path, 'r')
            except IOError:
                dson_file = open(dson_path, 'w')
                dson_file.write("[]")
                dson_file.close()
                dson_file = open(dson_path, 'r')

            data = dson.load(dson_file)
            try:
                elements = dict((element_data[cls.pk_field], element_data)
                                for element_data in data)
            except KeyError:
                msg = "Primary key was not found. Check that you have "
                msg += "configured the class correctly. In case you "
                msg += "have check your data source"
                raise AttributeError(msg)
        else:
            msg = "In order to use dson sources you should install "
            msg += " the 'dogeon' package"
            raise Exception(msg)

        return elements

    def write_elements(self, filepath, data):
        data_set = open('%s.dson' % filepath, 'w')
        dson_data = dson.dumps(data, indent=4)
        data_set.write(dson_data)
        data_set.close()
